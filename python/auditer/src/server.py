from openpyxl import load_workbook, Workbook
from openpyxl.writer.excel import save_virtual_workbook, ExcelWriter
from openpyxl.worksheet.table import Table as Otable
from openpyxl.worksheet.table import TableStyleInfo
from openpyxl.utils import get_column_letter
from flask import Flask, make_response, request, render_template, redirect, url_for, send_file, abort
from io import BytesIO
import sqlalchemy
from sqlalchemy import create_engine
#from sqlalchemy import Table, Column, Integer, String, MetaData, ForeignKey
from sqlalchemy import inspect
from zipfile import ZipFile,ZIP_DEFLATED
from dotenv import load_dotenv
from os import getenv

def save_virtual_workbook(workbook,):
    """Return an in-memory workbook, suitable for a Django response."""
    temp_buffer = BytesIO()
    archive = ZipFile(temp_buffer, 'w', ZIP_DEFLATED, allowZip64=True)

    writer = ExcelWriter(workbook, archive)

    try:
        writer.write_data()
    finally:
        archive.close()

    virtual_workbook = temp_buffer.getvalue()
    temp_buffer.close()
    return virtual_workbook


load_dotenv()
URI=getenv("URI")
server = Flask(__name__)

@server.route("/")
def hello():
    return "Hello World toto !"

@server.route('/sql')
@server.route('/sql/')
@server.route('/sql/<type_out>')
def employees(type_out='json'):
    content=""
    engine = create_engine(URI)
    inspector = inspect(engine)

    with engine.connect() as con:
        rs = con.execute('select first_name, last_name from employees limit 20')
        for h in rs._metadata.keys:
            content = f"{content}{h} / "
        content = f"{content}<hr><pre>"
        for r in rs:
            content = f"{content}* {r['first_name']} {r['last_name']}\n"
    content = f"{content}</pre>"

    response = make_response(content, 200)
    response.mimetype = "text/html"
    return response


@server.route("/xls")
def xlsgen():
    if request.args.get('token') != getenv('WEB_TOKEN'):
        abort(403)
    engine = create_engine(URI)
    inspector = inspect(engine)

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title=getenv('WS_TITLE'))

    with engine.connect() as con:
        rs = con.execute(getenv('SQL_REQUEST'))
        #ws.append(rs._metadata.keys)

        headers=[]
        for h in rs._metadata.keys:
            headers.append(h)
        ws.append(headers)

        for r in rs:
            rv=[]
            for col in headers:
                rv.append(r[col])
            ws.append(rv)

    table = Otable(displayName="Employees", ref="A1:" + get_column_letter(ws.max_column) + str(ws.max_row))

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleDark11", showFirstColumn=True, showLastColumn=True, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

    content = save_virtual_workbook(wb)
    resp = make_response(content)
    resp.headers['Content-Disposition'] = 'attachment; filename='+ getenv('WS_TITLE') +'.xlsx'
    resp.headers['Content-Type'] = 'application/x-xlsx'
    return resp

if __name__ == "__main__":
    server.run(host='0.0.0.0', debug=True)

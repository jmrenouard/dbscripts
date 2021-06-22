from openpyxl import load_workbook, Workbook
from openpyxl.writer.excel import save_virtual_workbook, ExcelWriter
from flask import Flask, make_response, request, render_template, redirect, url_for, send_file
from io import BytesIO

import mysql.connector

from zipfile import ZipFile,ZIP_DEFLATED

URI="mysql://root:secret@192.168.0.50:3306/employees"
def save_virtual_workbook(workbook,):
    """Return an in-memory workbook, suitable for a Django response."""
    temp_buffer = BytesIO()
    archive = ZipFile(temp_buffer, 'w', ZIP_DEFLATED, allowZip64=True)

    writer = ExcelWriter(workbook, archive)

    try:
        writer.write_data()
    finally:
        archive.close()

    virtual_workbook = temp_buffer.getvalue()
    temp_buffer.close()
    return virtual_workbook


server = Flask(__name__)


@server.route("/")
def hello():
    return "Hello World!"

@server.route('/sql')
def employees():
    content=""
    db = records.Database(URI)
    rows = db.query('select first_name, last_name from employees')

    response = make_response(content, 200)
    response.mimetype = "text/plain"
    return response

@server.route("/xls")
def xlsgen():
    wb = Workbook()
    ws = wb.create_sheet(title='test1')
    ws['A1'] = 'This is a test'

    db = records.Database(URI)
    rows = db.query('select first_name, last_name from employees')

    content = save_virtual_workbook(wb)
    resp = make_response(content)
    resp.headers['Content-Disposition'] = 'attachment; filename=test.xlsx'
    resp.headers['Content-Type'] = 'application/x-xlsx'
    return resp


if __name__ == "__main__":
    server.run(host='0.0.0.0')
